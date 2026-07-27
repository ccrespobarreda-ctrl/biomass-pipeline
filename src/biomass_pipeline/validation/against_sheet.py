"""Contraste de una extraccion contra el sheet historico.

Como el Excel ya trae el dato correcto, esto detecta al instante si un extractor
acierta o si un proveedor cambio el formato. Se usa en desarrollo y como red de
seguridad; en produccion, la deteccion la hacen las puertas de calidad + alertas.
"""
import datetime

import openpyxl


def cargar_hoja(ruta_xlsx: str, hoja: str):
    return openpyxl.load_workbook(ruta_xlsx, data_only=True)[hoja]


def fila_por_fecha(ws, fecha: datetime.date, col_fecha: int = 2) -> int | None:
    """Localiza la fila cuyo Date 2 (col B por defecto) coincide con 'fecha'."""
    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, col_fecha).value
        if isinstance(v, (datetime.datetime, datetime.date)):
            if (v.date() if isinstance(v, datetime.datetime) else v) == fecha:
                return r
    return None


def comparar(extraido: float, esperado, tol: float = 0.01) -> bool:
    try:
        return abs(float(extraido) - float(esperado)) < tol
    except (TypeError, ValueError):
        return False
